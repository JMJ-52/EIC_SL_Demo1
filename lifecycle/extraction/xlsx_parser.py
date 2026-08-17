from __future__ import annotations

import os

import openpyxl

from lifecycle.extraction.models import ParsedUnit


def parse_xlsx(path: str) -> list[ParsedUnit]:
    source_doc = os.path.basename(path)
    workbook = openpyxl.load_workbook(path, data_only=True)
    units = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        lines = []
        for row in worksheet.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            if any(cell.strip() for cell in cells):
                lines.append(" | ".join(cells))
        units.append(ParsedUnit(source_doc=source_doc, unit_label=sheet_name, text="\n".join(lines), images=[]))
    return units

